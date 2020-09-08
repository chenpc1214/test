import openpyxl as xl

file_place = "C:\\Users\\user\\Desktop\\Python 3.7\\python x excel\\transactions.xlsx"
wb = xl.load_workbook(file_place)
sheet = wb["工作表1"] #wb就是開啟的檔案，而在這的sheet就是開啟檔案中，你所選定工表格
cell = sheet["a1"]#所在表格頁中，你所選定的[欄]為名稱
cell = sheet.cell(1,1)
print(cell.value)#這是選定欄為的內容名稱
print(sheet.max_row)#這是欄(直的)位在Excel的總欄數

for row in range(1,sheet.max_row+1):#用for方法算出第一欄的欄數
    print(row)

for row in range(2,sheet.max_row+1):#用for方法找出第三欄的值
    cell_2 = sheet.cell(row,3)
    print(cell_2.value)

for row in range(2,sheet.max_row+1):#算出第四欄的欄位數值並另存新黨
    cell = sheet.cell(row,3)
    corrected_price = cell.value*0.9
    correct_price_cell = sheet.cell(row,4)
    correct_price_cell.value = corrected_price

    wb.save("transactions2.xlsx")
