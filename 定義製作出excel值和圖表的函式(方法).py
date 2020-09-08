import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["工作表1"] #wb就是開啟的檔案，而在這的sheet就是開啟檔案中，你所選定工表格

    print(cell.value)#這是選定欄為的內容名稱
    print(sheet.max_row)#這是欄(直的)位在Excel的總欄數

    for row in range(1,sheet.max_row+1):#用for方法算出第一欄的欄數
        print(row)

    for row in range(2,sheet.max_row+1):#用for方法找出第三欄的值
        cell_2 = sheet.cell(row,3)
        print(cell_2.value)

    for row in range(2,sheet.max_row+1):#算出第四欄的欄位數值並另存新黨
        cell = sheet.cell(row,3)
        corrected_price = cell.value*0.9
        correct_price_cell = sheet.cell(row,4)
        correct_price_cell.value = corrected_price

    values = Reference(sheet,#給出做圖表的參照範圍
              min_row = 2,
              max_row = sheet.max.row,
              min_col = 4,
              max_col = 4)

    chart = BarChart()#開始做出圖表
    chart.add_data(values)
    sheet.add_chart(chart,"E2")#在這要給出想要做出圖表的位置

    wb.save(filename)

